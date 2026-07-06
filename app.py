"""
אפליקציית ניהול דרישות
=======================
ניהול דרישות, תתי-דרישות, שיוך למודולים וזיהוי מי הציע כל דרישה.
ניהול משתמשים בידי האדמין בלבד.

הרצה מקומית:  python app.py
פריסה (Railway):  gunicorn app:app
"""

import os
import io
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, abort, send_file
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash

# ---------------------------------------------------------------------------
# הגדרות בסיס
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

# בסיס נתונים: Postgres ב-Railway אם קיים DATABASE_URL, אחרת SQLite מקומי.
db_url = os.environ.get("DATABASE_URL", "sqlite:///requirements.db")
# Railway מספק לעיתים postgres:// — SQLAlchemy דורש postgresql://
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 15 * 1024 * 1024  # מגבלת העלאה: 15MB לקובץ

db = SQLAlchemy(app)

# סטטוס אחיד לדרישות, משימות ובאגים
STATUSES = ["חדש", "בתהליך", "בוצע"]
PRIORITIES = ["נמוכה", "בינונית", "גבוהה", "קריטית"]

# תווית גרסה — לבדיקה שהפריסה התעדכנה
APP_VERSION = "גרסה 3.13 · תרשים מודולים ומימוש בפועל"


# ---------------------------------------------------------------------------
# מודלים
# ---------------------------------------------------------------------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    display_name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(40), default="")  # לטובת התראות וואטסאפ בעתיד
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    log_seen_at = db.Column(db.DateTime, nullable=True)  # מתי המשתמש צפה בלוג לאחרונה
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    proposed = db.relationship(
        "Requirement", backref="proposer",
        foreign_keys="Requirement.proposer_id"
    )

    def set_password(self, raw):
        self.password_hash = generate_password_hash(raw)

    def check_password(self, raw):
        return check_password_hash(self.password_hash, raw)


class Module(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    color = db.Column(db.String(7), default="#6b7280")  # צבע תווית
    parent_id = db.Column(db.Integer, db.ForeignKey("module.id"), nullable=True)  # מודול-אב
    requirements = db.relationship("Requirement", backref="module")
    children = db.relationship(
        "Module",
        backref=db.backref("parent", remote_side=[id]),
        foreign_keys=[parent_id],
    )


class ModuleLink(db.Model):
    """קשר בין שני מודולים (כולל תתי-מודולים), עם תיאור אופציונלי."""
    id = db.Column(db.Integer, primary_key=True)
    source_id = db.Column(db.Integer, db.ForeignKey("module.id"), nullable=False)
    target_id = db.Column(db.Integer, db.ForeignKey("module.id"), nullable=False)
    label = db.Column(db.String(200), default="")
    source = db.relationship("Module", foreign_keys=[source_id],
                             backref=db.backref("links_out", cascade="all, delete-orphan"))
    target = db.relationship("Module", foreign_keys=[target_id])


class Requirement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(300), nullable=False)
    description = db.Column(db.Text, default="")
    status = db.Column(db.String(40), default="חדש")
    priority = db.Column(db.String(40), default="בינונית")

    module_id = db.Column(db.Integer, db.ForeignKey("module.id"), nullable=True)
    proposer_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    implementer_id = db.Column(db.Integer, nullable=True)      # מוקצה לביצוע (מזהה)
    implementer_name = db.Column(db.String(120), default="")   # מוקצה לביצוע (שם)
    actual_implementer_id = db.Column(db.Integer, nullable=True)     # מי מימש בפועל (מזהה)
    actual_implementer_name = db.Column(db.String(120), default="")  # מי מימש בפועל (שם)
    parent_id = db.Column(db.Integer, db.ForeignKey("requirement.id"), nullable=True)

    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # תתי-דרישות
    children = db.relationship(
        "Requirement",
        backref=db.backref("parent", remote_side=[id]),
        cascade="all, delete-orphan",
        foreign_keys=[parent_id],
    )
    creator = db.relationship("User", foreign_keys=[created_by_id])


class ActionLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, nullable=True)        # מי ביצע
    user_name = db.Column(db.String(120), default="")     # שם משוכפל (נשמר גם אם המשתמש נמחק)
    action = db.Column(db.String(80), default="")         # סוג הפעולה
    target = db.Column(db.String(300), default="")        # על מה בוצעה
    details = db.Column(db.String(500), default="")       # פירוט נוסף


class Comment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    requirement_id = db.Column(db.Integer, db.ForeignKey("requirement.id"), nullable=False)
    author_id = db.Column(db.Integer, nullable=True)
    author_name = db.Column(db.String(120), default="")
    body = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    requirement = db.relationship(
        "Requirement",
        backref=db.backref("comments", cascade="all, delete-orphan", order_by="Comment.created_at"),
    )


class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)  # הנמען
    requirement_id = db.Column(db.Integer, db.ForeignKey("requirement.id"), nullable=True)
    bug_id = db.Column(db.Integer, db.ForeignKey("bug.id"), nullable=True)
    text = db.Column(db.String(400), default="")
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Bug(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(300), nullable=False)          # שם הבאג
    description = db.Column(db.Text, default="")               # תיאור הבאג
    solution = db.Column(db.Text, default="")                  # רעיון לפתרון
    requirement_id = db.Column(db.Integer, db.ForeignKey("requirement.id"), nullable=True)  # קישור לדרישה
    status = db.Column(db.String(20), default="חדש")           # חדש / בתהליך / בוצע
    priority = db.Column(db.String(40), default="בינונית")     # עדיפות

    assignee_id = db.Column(db.Integer, nullable=True)         # אחראי על הבאג (מזהה)
    assignee_name = db.Column(db.String(120), default="")      # אחראי (שם לתצוגה)

    opened_by_id = db.Column(db.Integer, nullable=True)
    opened_by_name = db.Column(db.String(120), default="")     # מי פתח
    created_at = db.Column(db.DateTime, default=datetime.utcnow)  # מתי נפתח

    status_changed_by = db.Column(db.String(120), default="")  # מי שינה סטטוס לאחרונה
    status_changed_at = db.Column(db.DateTime, nullable=True)

    requirement = db.relationship("Requirement")


class BugComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bug_id = db.Column(db.Integer, db.ForeignKey("bug.id"), nullable=False)
    author_id = db.Column(db.Integer, nullable=True)
    author_name = db.Column(db.String(120), default="")
    body = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    bug = db.relationship(
        "Bug",
        backref=db.backref("comments", cascade="all, delete-orphan", order_by="BugComment.created_at"),
    )


BUG_STATUSES = STATUSES


class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(300), nullable=False)          # שם המשימה
    description = db.Column(db.Text, default="")               # תיאור
    assignee_id = db.Column(db.Integer, nullable=True)         # אחראי (מזהה)
    assignee_name = db.Column(db.String(120), default="")      # אחראי (שם לתצוגה)
    due_date = db.Column(db.Date, nullable=True)               # תאריך יעד
    status = db.Column(db.String(20), default="חדש")           # חדש / בתהליך / בוצע
    priority = db.Column(db.String(40), default="בינונית")     # עדיפות

    created_by_id = db.Column(db.Integer, nullable=True)
    created_by_name = db.Column(db.String(120), default="")    # מי פתח
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


TASK_STATUSES = STATUSES


class Attachment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    requirement_id = db.Column(db.Integer, db.ForeignKey("requirement.id"), nullable=True)
    bug_id = db.Column(db.Integer, db.ForeignKey("bug.id"), nullable=True)
    filename = db.Column(db.String(300), default="")
    mimetype = db.Column(db.String(120), default="")
    data = db.Column(db.LargeBinary, nullable=False)   # תוכן הקובץ נשמר ב-DB
    uploaded_by = db.Column(db.String(120), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    requirement = db.relationship(
        "Requirement",
        backref=db.backref("attachments", cascade="all, delete-orphan"),
    )
    bug = db.relationship(
        "Bug",
        backref=db.backref("attachments", cascade="all, delete-orphan"),
    )

    @property
    def is_image(self):
        return (self.mimetype or "").startswith("image/")


def send_whatsapp(phone, message):
    """
    נקודת חיבור עתידית לוואטסאפ.
    כיום אינה פעילה. כדי להפעיל בעתיד דרך Twilio:
      1. להתקין: pip install twilio (ולהוסיף ל-requirements.txt)
      2. להגדיר משתני סביבה: TWILIO_SID, TWILIO_TOKEN, TWILIO_WHATSAPP_FROM
      3. לממש כאן את השליחה בפועל.
    מוחזר True אם נשלח, אחרת False.
    """
    if not phone:
        return False
    # --- מקום למימוש Twilio בעתיד ---
    return False


def modules_for_select():
    """רשימת מודולים שטוחה בסדר היררכי: ראשי ואחריו תתי-המודולים שלו,
    עם שם תצוגה 'אב ← תת' לתתי-מודולים."""
    result = []
    for m in Module.query.filter_by(parent_id=None).order_by(Module.name).all():
        m.select_label = m.name
        result.append(m)
        for c in sorted(m.children, key=lambda x: x.name):
            c.select_label = f"{m.name} ← {c.name}"
            result.append(c)
    return result


def log_action(action, target="", details=""):
    """רישום פעולה ללוג. נקרא בתוך נתיב, אחרי שיש משתמש מחובר."""
    u = current_user()
    entry = ActionLog(
        user_id=u.id if u else None,
        user_name=u.display_name if u else "אורח",
        action=action,
        target=str(target)[:300],
        details=str(details)[:500],
    )
    db.session.add(entry)
    db.session.commit()


def parse_mentions(body, exclude_id=None):
    """
    מחזירה רשימת מזהי משתמשים שאוזכרו בטקסט בעזרת @שם.
    תואמת שמות מלאים (כולל רווחים). שמות ארוכים נבדקים קודם כדי
    שאזכור של '@יוסי כהן' לא ייתפס בטעות גם כ'@יוסי'.
    """
    if not body or "@" not in body:
        return []
    work = body
    ids = []
    users = sorted(User.query.all(), key=lambda u: len(u.display_name), reverse=True)
    for u in users:
        token = "@" + u.display_name
        if token in work:
            work = work.replace(token, " ")
            if not (exclude_id and u.id == exclude_id):
                ids.append(u.id)
    return ids


# ---------------------------------------------------------------------------
# אימות והרשאות
# ---------------------------------------------------------------------------
def current_user():
    uid = session.get("user_id")
    if uid:
        return db.session.get(User, uid)
    return None


@app.context_processor
def inject_user():
    u = current_user()
    unread = 0
    nav = {"req": 0, "bug": 0, "task": 0, "logs_new": False}
    if u:
        unread = Notification.query.filter_by(user_id=u.id, is_read=False).count()
        # מונה פריטים שלא בוצעו לכל טאב
        nav["req"] = Requirement.query.filter(
            Requirement.parent_id.is_(None), Requirement.status != "בוצע"
        ).count()
        nav["bug"] = Bug.query.filter(Bug.status != "בוצע").count()
        nav["task"] = Task.query.filter(Task.status != "בוצע").count()
        # אינדיקציית לוג חדש (לאדמין בלבד): רשומות חדשות ממשתמש אחר
        if u.is_admin:
            q = ActionLog.query.filter(ActionLog.user_id != u.id)
            if u.log_seen_at:
                q = q.filter(ActionLog.timestamp > u.log_seen_at)
            nav["logs_new"] = db.session.query(q.exists()).scalar()
    return {"current_user": u, "unread_count": unread, "app_version": APP_VERSION, "nav": nav}


@app.template_filter("israel_time")
def israel_time(dt):
    """המרת זמן UTC לשעון ישראל לתצוגה."""
    if dt is None:
        return ""
    try:
        from zoneinfo import ZoneInfo
        return dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(
            ZoneInfo("Asia/Jerusalem")
        ).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return dt.strftime("%d/%m/%Y %H:%M")


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u:
            return redirect(url_for("login"))
        if not u.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return wrapper


# ---------------------------------------------------------------------------
# נתיבי אימות
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session["user_id"] = user.id
            log_action("כניסה למערכת")
            return redirect(url_for("index"))
        flash("שם משתמש או סיסמה שגויים", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    log_action("יציאה מהמערכת")
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# נתיב ראשי — רשימת הדרישות
# ---------------------------------------------------------------------------
@app.route("/")
@login_required
def index():
    module_filter = request.args.get("module", type=int)
    status_filter = request.args.get("status", "")
    proposer_filter = request.args.get("proposer", type=int)
    implementer_filter = request.args.get("implementer", type=int)
    priority_filter = request.args.get("priority", "")

    query = Requirement.query.filter_by(parent_id=None)
    if module_filter:
        query = query.filter_by(module_id=module_filter)
    if status_filter:
        query = query.filter_by(status=status_filter)
    if proposer_filter:
        query = query.filter_by(proposer_id=proposer_filter)
    if implementer_filter:
        query = query.filter_by(implementer_id=implementer_filter)
    if priority_filter:
        query = query.filter_by(priority=priority_filter)

    requirements = query.order_by(Requirement.created_at.desc()).all()
    modules = modules_for_select()
    users = User.query.order_by(User.display_name).all()
    return render_template(
        "index.html",
        requirements=requirements,
        modules=modules,
        users=users,
        statuses=STATUSES,
        priorities=PRIORITIES,
        module_filter=module_filter,
        status_filter=status_filter,
        proposer_filter=proposer_filter,
        implementer_filter=implementer_filter,
        priority_filter=priority_filter,
    )


# ---------------------------------------------------------------------------
# ייצוא וייבוא דרישות — Excel
# ---------------------------------------------------------------------------
EXCEL_HEADERS = ["מזהה", "כותרת", "תיאור", "סטטוס", "עדיפות", "מודול",
                 "מי הציע", "מי מימש", "דרישת אב (מזהה)", "נוצרה בתאריך"]


@app.route("/requirements/export")
@login_required
def requirements_export():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "דרישות"
    ws.sheet_view.rightToLeft = True
    ws.append(EXCEL_HEADERS)
    # כל הדרישות: אבות קודם ואז ילדים, כדי שהייבוא יוכל לקשר
    all_reqs = Requirement.query.order_by(
        Requirement.parent_id.isnot(None), Requirement.id
    ).all()
    for r in all_reqs:
        ws.append([
            r.id,
            r.title,
            r.description or "",
            r.status or "",
            r.priority or "",
            r.module.name if r.module else "",
            r.proposer.display_name if r.proposer else "",
            r.implementer_name or "",
            r.parent_id or "",
            r.created_at.strftime("%d/%m/%Y") if r.created_at else "",
        ])
    # רוחב עמודות סביר
    for col, w in zip("ABCDEFGHIJ", (8, 40, 50, 12, 12, 18, 18, 18, 16, 14)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action("ייצוא דרישות לאקסל", f"{len(all_reqs)} דרישות")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="requirements.xlsx",
    )


@app.route("/requirements/import", methods=["POST"])
@login_required
def requirements_import():
    from openpyxl import load_workbook
    f = request.files.get("file")
    if not f or not f.filename:
        flash("לא נבחר קובץ", "error")
        return redirect(url_for("index"))
    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
    except Exception:
        flash("הקובץ אינו קובץ אקסל תקין (.xlsx)", "error")
        return redirect(url_for("index"))

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    me = current_user()
    users_by_name = {u.display_name: u for u in User.query.all()}
    modules_by_name = {m.name: m for m in Module.query.all()}

    id_map = {}   # מזהה מהקובץ -> דרישה חדשה (לקישור אב-ילד)
    created = 0
    skipped = 0
    # שני מעברים: קודם דרישות-אב (ללא אב), ואז ילדים
    def row_get(row, i):
        return (str(row[i]).strip() if i < len(row) and row[i] is not None else "")

    parsed = []
    for row in rows:
        title = row_get(row, 1)
        if not title:
            skipped += 1
            continue
        parsed.append({
            "file_id": row_get(row, 0),
            "title": title,
            "description": row_get(row, 2),
            "status": row_get(row, 3) if row_get(row, 3) in STATUSES else "חדש",
            "priority": row_get(row, 4) if row_get(row, 4) in PRIORITIES else "בינונית",
            "module": row_get(row, 5),
            "proposer": row_get(row, 6),
            "implementer": row_get(row, 7),
            "parent_file_id": row_get(row, 8),
        })

    for phase_parents in (True, False):
        for p in parsed:
            is_parent = not p["parent_file_id"]
            if is_parent != phase_parents:
                continue
            module = modules_by_name.get(p["module"])
            # מודול חדש נוצר אוטומטית אם לא קיים
            if p["module"] and not module:
                module = Module(name=p["module"])
                db.session.add(module)
                db.session.flush()
                modules_by_name[p["module"]] = module
            proposer = users_by_name.get(p["proposer"])
            implementer = users_by_name.get(p["implementer"])
            parent_id = None
            if p["parent_file_id"]:
                parent = id_map.get(p["parent_file_id"])
                parent_id = parent.id if parent else None
            req = Requirement(
                title=p["title"],
                description=p["description"],
                status=p["status"],
                priority=p["priority"],
                module_id=module.id if module else None,
                proposer_id=proposer.id if proposer else None,
                implementer_id=implementer.id if implementer else None,
                implementer_name=implementer.display_name if implementer else (p["implementer"] or ""),
                parent_id=parent_id,
                created_by_id=me.id,
            )
            db.session.add(req)
            db.session.flush()
            if p["file_id"]:
                id_map[p["file_id"]] = req
            created += 1
    db.session.commit()
    log_action("ייבוא דרישות מאקסל", f"{created} נוספו, {skipped} דולגו")
    flash(f"ייבוא הושלם: {created} דרישות נוספו" + (f", {skipped} שורות דולגו (ללא כותרת)" if skipped else ""), "ok")
    return redirect(url_for("index"))


@app.route("/requirement/<int:req_id>")
@login_required
def requirement_detail(req_id):
    req = db.session.get(Requirement, req_id)
    if not req:
        abort(404)
    modules = modules_for_select()
    users = User.query.order_by(User.display_name).all()
    return render_template(
        "requirement.html",
        req=req,
        modules=modules,
        users=users,
        statuses=STATUSES,
        priorities=PRIORITIES,
    )


@app.route("/requirement/add", methods=["POST"])
@login_required
def requirement_add():
    title = request.form.get("title", "").strip()
    if not title:
        flash("חובה להזין כותרת לדרישה", "error")
        return redirect(request.referrer or url_for("index"))

    parent_id = request.form.get("parent_id", type=int)
    module_id = request.form.get("module_id", type=int) or None
    proposer_id = request.form.get("proposer_id", type=int) or None
    implementer_id = request.form.get("implementer_id", type=int) or None
    implementer_name = ""
    if implementer_id:
        imp = db.session.get(User, implementer_id)
        implementer_name = imp.display_name if imp else ""

    req = Requirement(
        title=title,
        description=request.form.get("description", "").strip(),
        status=request.form.get("status") or "חדש",
        priority=request.form.get("priority") or "בינונית",
        module_id=module_id,
        proposer_id=proposer_id,
        implementer_id=implementer_id,
        implementer_name=implementer_name,
        parent_id=parent_id,
        created_by_id=current_user().id,
    )
    db.session.add(req)
    db.session.commit()
    log_action("הוספת תת-דרישה" if parent_id else "הוספת דרישה", title)
    flash("הדרישה נוספה בהצלחה", "ok")
    if parent_id:
        return redirect(url_for("requirement_detail", req_id=parent_id))
    return redirect(url_for("index"))


@app.route("/requirement/<int:req_id>/edit", methods=["POST"])
@login_required
def requirement_edit(req_id):
    req = db.session.get(Requirement, req_id)
    if not req:
        abort(404)
    req.title = request.form.get("title", req.title).strip()
    req.description = request.form.get("description", "").strip()
    req.status = request.form.get("status") or req.status
    req.priority = request.form.get("priority") or req.priority
    req.module_id = request.form.get("module_id", type=int) or None
    req.proposer_id = request.form.get("proposer_id", type=int) or None
    prev_impl = req.implementer_id
    implementer_id = request.form.get("implementer_id", type=int) or None
    req.implementer_id = implementer_id
    if implementer_id:
        imp = db.session.get(User, implementer_id)
        req.implementer_name = imp.display_name if imp else ""
    else:
        req.implementer_name = ""
    # מי מימש בפועל — רלוונטי כשהדרישה בסטטוס 'בוצע'.
    # אם נסגרה בלי בחירה מפורשת, ברירת המחדל היא המוקצה לביצוע.
    if req.status == "בוצע":
        actual_id = request.form.get("actual_implementer_id", type=int) or None
        if actual_id:
            au = db.session.get(User, actual_id)
            req.actual_implementer_id = actual_id
            req.actual_implementer_name = au.display_name if au else ""
        elif not req.actual_implementer_id and req.implementer_id:
            req.actual_implementer_id = req.implementer_id
            req.actual_implementer_name = req.implementer_name
    else:
        # דרישה שנפתחה מחדש — מנקים את המימוש בפועל
        req.actual_implementer_id = None
        req.actual_implementer_name = ""
    db.session.commit()
    # התראה לממש אם שויך ושונה מקודם
    me = current_user()
    if implementer_id and implementer_id != prev_impl and implementer_id != me.id:
        target = db.session.get(User, implementer_id)
        if target:
            db.session.add(Notification(
                user_id=target.id, requirement_id=req.id,
                text=f"{me.display_name} שייך/ה אותך כמממש/ת של \"{req.title}\"",
            ))
            send_whatsapp(target.phone, f"שויכת למימוש דרישה: {req.title}")
            db.session.commit()
    log_action("עריכת דרישה", req.title, f"סטטוס: {req.status}")
    flash("הדרישה עודכנה", "ok")
    return redirect(request.referrer or url_for("requirement_detail", req_id=req_id))


@app.route("/requirement/<int:req_id>/delete", methods=["POST"])
@login_required
def requirement_delete(req_id):
    req = db.session.get(Requirement, req_id)
    if not req:
        abort(404)
    parent_id = req.parent_id
    req_title = req.title

    # איסוף הדרישה וכל צאצאיה (תתי-דרישות בכל עומק)
    ids = []
    stack = [req]
    while stack:
        node = stack.pop()
        ids.append(node.id)
        stack.extend(node.children)

    # ניתוק רשומות שמפנות לדרישות (PostgreSQL אוכף מפתחות זרים):
    # באגים מקושרים — נשארים, רק הקישור מתנתק; התראות על הדרישות — נמחקות.
    try:
        Bug.query.filter(Bug.requirement_id.in_(ids)).update(
            {"requirement_id": None}, synchronize_session=False)
        Notification.query.filter(Notification.requirement_id.in_(ids)).delete(
            synchronize_session=False)
        db.session.delete(req)  # תגובות וקבצים נמחקים בקסקדה
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"המחיקה נכשלה: {e}", "error")
        return redirect(url_for("requirement_detail", req_id=req_id))

    log_action("מחיקת דרישה", req_title)
    flash("הדרישה נמחקה", "ok")
    if parent_id:
        return redirect(url_for("requirement_detail", req_id=parent_id))
    return redirect(url_for("index"))


@app.route("/requirement/<int:req_id>/comment", methods=["POST"])
@login_required
def comment_add(req_id):
    req = db.session.get(Requirement, req_id)
    if not req:
        abort(404)
    body = request.form.get("body", "").strip()
    if not body:
        flash("לא ניתן לשלוח תגובה ריקה", "error")
        return redirect(url_for("requirement_detail", req_id=req_id))

    me = current_user()
    comment = Comment(
        requirement_id=req.id,
        author_id=me.id,
        author_name=me.display_name,
        body=body,
    )
    db.session.add(comment)
    db.session.commit()
    log_action("תגובה בצ'אט", req.title)

    # אזכורים בטקסט (@שם) → התראות
    me_id = me.id
    for tid in parse_mentions(body, exclude_id=me_id):
        target_user = db.session.get(User, tid)
        if not target_user:
            continue
        note = Notification(
            user_id=target_user.id,
            requirement_id=req.id,
            text=f"{me.display_name} אזכר/ה אותך בצ'אט של \"{req.title}\"",
        )
        db.session.add(note)
        # תשתית עתידית לוואטסאפ (כיום אינה פעילה)
        send_whatsapp(target_user.phone, f"אוזכרת בדרישה: {req.title}")
    db.session.commit()
    flash("התגובה נוספה", "ok")
    return redirect(url_for("requirement_detail", req_id=req_id) + "#chat")


# ---------------------------------------------------------------------------
# התראות
# ---------------------------------------------------------------------------
@app.route("/notifications")
@login_required
def notifications():
    me = current_user()
    notes = (
        Notification.query.filter_by(user_id=me.id)
        .order_by(Notification.created_at.desc())
        .limit(100)
        .all()
    )
    # סימון הכל כנקרא לאחר הצפייה
    Notification.query.filter_by(user_id=me.id, is_read=False).update({"is_read": True})
    db.session.commit()
    return render_template("notifications.html", notes=notes)


# ---------------------------------------------------------------------------
# חיפוש כללי — דרישות, משימות ובאגים יחד
# ---------------------------------------------------------------------------
@app.route("/search")
@login_required
def search():
    q = request.args.get("q", "").strip()
    type_f = request.args.get("type", "")       # '' / requirement / task / bug
    status_f = request.args.get("status", "")
    priority_f = request.args.get("priority", "")
    assignee_f = request.args.get("assignee", type=int)
    results = []
    like = f"%{q}%"

    if type_f in ("", "requirement"):
        qr = Requirement.query
        if q:
            qr = qr.filter(db.or_(Requirement.title.ilike(like), Requirement.description.ilike(like)))
        if status_f:
            qr = qr.filter_by(status=status_f)
        if priority_f:
            qr = qr.filter_by(priority=priority_f)
        if assignee_f:
            qr = qr.filter_by(implementer_id=assignee_f)
        for r in qr.all():
            results.append({
                "type": "דרישה", "title": r.title, "description": r.description,
                "status": r.status, "priority": r.priority,
                "who": r.implementer_name or "", "url": url_for("requirement_detail", req_id=r.id),
            })

    if type_f in ("", "task"):
        qt = Task.query
        if q:
            qt = qt.filter(db.or_(Task.title.ilike(like), Task.description.ilike(like)))
        if status_f:
            qt = qt.filter_by(status=status_f)
        if priority_f:
            qt = qt.filter_by(priority=priority_f)
        if assignee_f:
            qt = qt.filter_by(assignee_id=assignee_f)
        for t in qt.all():
            results.append({
                "type": "משימה", "title": t.title, "description": t.description,
                "status": t.status, "priority": t.priority,
                "who": t.assignee_name or "", "url": url_for("tasks"),
            })

    if type_f in ("", "bug"):
        qb = Bug.query
        if q:
            qb = qb.filter(db.or_(Bug.title.ilike(like), Bug.description.ilike(like)))
        if status_f:
            qb = qb.filter_by(status=status_f)
        if priority_f:
            qb = qb.filter_by(priority=priority_f)
        if assignee_f:
            qb = qb.filter_by(assignee_id=assignee_f)
        for b in qb.all():
            results.append({
                "type": "באג", "title": b.title, "description": b.description,
                "status": b.status, "priority": b.priority,
                "who": b.assignee_name or "", "url": url_for("bug_detail", bug_id=b.id),
            })

    users = User.query.order_by(User.display_name).all()
    all_statuses = STATUSES
    searched = bool(q or type_f or status_f or priority_f or assignee_f)
    return render_template(
        "search.html", results=results, users=users,
        priorities=PRIORITIES, all_statuses=all_statuses, searched=searched,
        q=q, type_f=type_f, status_f=status_f, priority_f=priority_f, assignee_f=assignee_f,
    )


# ---------------------------------------------------------------------------
# ניהול באגים
# ---------------------------------------------------------------------------
@app.route("/bugs")
@login_required
def bugs():
    status_filter = request.args.get("status", "")
    assignee_filter = request.args.get("assignee", type=int)
    priority_filter = request.args.get("priority", "")
    query = Bug.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    if assignee_filter:
        query = query.filter_by(assignee_id=assignee_filter)
    if priority_filter:
        query = query.filter_by(priority=priority_filter)
    # פתוחים קודם, ואז לפי תאריך יורד
    bug_list = query.order_by(Bug.status.desc(), Bug.created_at.desc()).all()
    requirements = Requirement.query.order_by(Requirement.title).all()
    users = User.query.order_by(User.display_name).all()
    open_count = Bug.query.filter(Bug.status != "בוצע").count()
    return render_template(
        "bugs.html",
        bugs=bug_list,
        requirements=requirements,
        users=users,
        statuses=BUG_STATUSES,
        priorities=PRIORITIES,
        status_filter=status_filter,
        assignee_filter=assignee_filter,
        priority_filter=priority_filter,
        open_count=open_count,
    )


def _notify_assignee(bug, assignee_id, actor):
    """יצירת התראה למשתמש שהבאג הוקצה לו (אם אינו עצמו)."""
    if not assignee_id or assignee_id == actor.id:
        return
    target = db.session.get(User, assignee_id)
    if not target:
        return
    db.session.add(Notification(
        user_id=target.id,
        bug_id=bug.id,
        text=f"{actor.display_name} הקצה/תה לך את הבאג \"{bug.title}\"",
    ))
    send_whatsapp(target.phone, f"הוקצה לך באג: {bug.title}")


@app.route("/bugs/add", methods=["POST"])
@login_required
def bug_add():
    title = request.form.get("title", "").strip()
    if not title:
        flash("חובה להזין שם לבאג", "error")
        return redirect(url_for("bugs"))
    me = current_user()
    assignee_id = request.form.get("assignee_id", type=int) or None
    assignee_name = ""
    if assignee_id:
        a = db.session.get(User, assignee_id)
        assignee_name = a.display_name if a else ""
    bug = Bug(
        title=title,
        description=request.form.get("description", "").strip(),
        solution=request.form.get("solution", "").strip(),
        requirement_id=request.form.get("requirement_id", type=int) or None,
        status="חדש",
        priority=request.form.get("priority") or "בינונית",
        assignee_id=assignee_id,
        assignee_name=assignee_name,
        opened_by_id=me.id,
        opened_by_name=me.display_name,
    )
    db.session.add(bug)
    db.session.commit()
    _notify_assignee(bug, assignee_id, me)
    db.session.commit()
    log_action("פתיחת באג", title)
    flash("הבאג נפתח בהצלחה", "ok")
    return redirect(url_for("bugs"))


@app.route("/bug/<int:bug_id>")
@login_required
def bug_detail(bug_id):
    bug = db.session.get(Bug, bug_id)
    if not bug:
        abort(404)
    requirements = Requirement.query.order_by(Requirement.title).all()
    users = User.query.order_by(User.display_name).all()
    return render_template(
        "bug_detail.html", bug=bug, requirements=requirements,
        users=users, statuses=BUG_STATUSES, priorities=PRIORITIES,
    )


@app.route("/bug/<int:bug_id>/edit", methods=["POST"])
@login_required
def bug_edit(bug_id):
    bug = db.session.get(Bug, bug_id)
    if not bug:
        abort(404)
    me = current_user()
    prev_assignee = bug.assignee_id
    bug.title = request.form.get("title", bug.title).strip()
    bug.description = request.form.get("description", "").strip()
    bug.solution = request.form.get("solution", "").strip()
    bug.priority = request.form.get("priority") or bug.priority
    bug.requirement_id = request.form.get("requirement_id", type=int) or None
    assignee_id = request.form.get("assignee_id", type=int) or None
    bug.assignee_id = assignee_id
    if assignee_id:
        a = db.session.get(User, assignee_id)
        bug.assignee_name = a.display_name if a else ""
    else:
        bug.assignee_name = ""
    db.session.commit()
    # התראה רק אם האחראי השתנה
    if assignee_id and assignee_id != prev_assignee:
        _notify_assignee(bug, assignee_id, me)
        db.session.commit()
    log_action("עריכת באג", bug.title)
    flash("הבאג עודכן", "ok")
    return redirect(request.referrer or url_for("bugs"))


@app.route("/bug/<int:bug_id>/status", methods=["POST"])
@login_required
def bug_status(bug_id):
    bug = db.session.get(Bug, bug_id)
    if not bug:
        abort(404)
    new_status = request.form.get("status", "")
    if new_status in BUG_STATUSES and new_status != bug.status:
        me = current_user()
        bug.status = new_status
        bug.status_changed_by = me.display_name
        bug.status_changed_at = datetime.utcnow()
        db.session.commit()
        log_action("שינוי סטטוס באג", bug.title, f"ל-{new_status}")
        flash(f"סטטוס הבאג שונה ל-{new_status}", "ok")
    return redirect(request.referrer or url_for("bugs"))


@app.route("/bug/<int:bug_id>/comment", methods=["POST"])
@login_required
def bug_comment_add(bug_id):
    bug = db.session.get(Bug, bug_id)
    if not bug:
        abort(404)
    body = request.form.get("body", "").strip()
    if not body:
        flash("לא ניתן לשלוח תגובה ריקה", "error")
        return redirect(url_for("bug_detail", bug_id=bug_id))
    me = current_user()
    db.session.add(BugComment(
        bug_id=bug.id, author_id=me.id, author_name=me.display_name, body=body,
    ))
    db.session.commit()
    log_action("תגובה בצ'אט באג", bug.title)

    # אזכורים בטקסט (@שם) → התראות
    for tid in parse_mentions(body, exclude_id=me.id):
        target = db.session.get(User, tid)
        if not target:
            continue
        db.session.add(Notification(
            user_id=target.id, bug_id=bug.id,
            text=f"{me.display_name} אזכר/ה אותך בצ'אט של הבאג \"{bug.title}\"",
        ))
        send_whatsapp(target.phone, f"אוזכרת בבאג: {bug.title}")
    db.session.commit()
    flash("התגובה נוספה", "ok")
    return redirect(url_for("bug_detail", bug_id=bug_id) + "#chat")


@app.route("/bug/<int:bug_id>/delete", methods=["POST"])
@login_required
def bug_delete(bug_id):
    bug = db.session.get(Bug, bug_id)
    if bug:
        title = bug.title
        try:
            Notification.query.filter_by(bug_id=bug.id).delete(synchronize_session=False)
            db.session.delete(bug)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"המחיקה נכשלה: {e}", "error")
            return redirect(url_for("bug_detail", bug_id=bug_id))
        log_action("מחיקת באג", title)
        flash("הבאג נמחק", "ok")
    return redirect(url_for("bugs"))


# ---------------------------------------------------------------------------
# קבצים מצורפים (נשמרים במסד הנתונים)
# ---------------------------------------------------------------------------
def _save_upload(file_storage, requirement_id=None, bug_id=None):
    if not file_storage or not file_storage.filename:
        return False
    data = file_storage.read()
    if not data:
        return False
    att = Attachment(
        requirement_id=requirement_id,
        bug_id=bug_id,
        filename=file_storage.filename,
        mimetype=file_storage.mimetype or "application/octet-stream",
        data=data,
        uploaded_by=current_user().display_name,
    )
    db.session.add(att)
    db.session.commit()
    return True


@app.route("/requirement/<int:req_id>/attach", methods=["POST"])
@login_required
def requirement_attach(req_id):
    req = db.session.get(Requirement, req_id)
    if not req:
        abort(404)
    if _save_upload(request.files.get("file"), requirement_id=req.id):
        log_action("צירוף קובץ לדרישה", req.title)
        flash("הקובץ צורף", "ok")
    else:
        flash("לא נבחר קובץ", "error")
    return redirect(url_for("requirement_detail", req_id=req_id) + "#files")


@app.route("/bug/<int:bug_id>/attach", methods=["POST"])
@login_required
def bug_attach(bug_id):
    bug = db.session.get(Bug, bug_id)
    if not bug:
        abort(404)
    if _save_upload(request.files.get("file"), bug_id=bug.id):
        log_action("צירוף קובץ לבאג", bug.title)
        flash("הקובץ צורף", "ok")
    else:
        flash("לא נבחר קובץ", "error")
    return redirect(url_for("bug_detail", bug_id=bug_id) + "#files")


@app.route("/attachment/<int:att_id>")
@login_required
def attachment_view(att_id):
    att = db.session.get(Attachment, att_id)
    if not att:
        abort(404)
    # תמונות מוצגות בדפדפן, שאר הקבצים יורדים
    as_download = not att.is_image
    return send_file(
        io.BytesIO(att.data),
        mimetype=att.mimetype or "application/octet-stream",
        as_attachment=as_download,
        download_name=att.filename or f"attachment-{att.id}",
    )


@app.route("/attachment/<int:att_id>/delete", methods=["POST"])
@login_required
def attachment_delete(att_id):
    att = db.session.get(Attachment, att_id)
    if att:
        req_id, bug_id = att.requirement_id, att.bug_id
        db.session.delete(att)
        db.session.commit()
        log_action("מחיקת קובץ מצורף", att.filename)
        flash("הקובץ נמחק", "ok")
        if bug_id:
            return redirect(url_for("bug_detail", bug_id=bug_id) + "#files")
        if req_id:
            return redirect(url_for("requirement_detail", req_id=req_id) + "#files")
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# ניהול משימות
# ---------------------------------------------------------------------------
def _parse_date(value):
    """המרת מחרוזת תאריך מטופס (YYYY-MM-DD) לאובייקט תאריך."""
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


@app.route("/tasks")
@login_required
def tasks():
    status_filter = request.args.get("status", "")
    assignee_filter = request.args.get("assignee", type=int)
    priority_filter = request.args.get("priority", "")
    query = Task.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    if assignee_filter:
        query = query.filter_by(assignee_id=assignee_filter)
    if priority_filter:
        query = query.filter_by(priority=priority_filter)
    task_list = query.order_by(Task.status, Task.due_date.is_(None), Task.due_date).all()
    users = User.query.order_by(User.display_name).all()
    open_count = Task.query.filter(Task.status != "בוצע").count()
    return render_template(
        "tasks.html",
        tasks=task_list,
        users=users,
        statuses=TASK_STATUSES,
        priorities=PRIORITIES,
        status_filter=status_filter,
        assignee_filter=assignee_filter,
        priority_filter=priority_filter,
        open_count=open_count,
        today=date.today(),
    )


@app.route("/tasks/add", methods=["POST"])
@login_required
def task_add():
    title = request.form.get("title", "").strip()
    if not title:
        flash("חובה להזין שם למשימה", "error")
        return redirect(url_for("tasks"))
    me = current_user()
    assignee_id = request.form.get("assignee_id", type=int) or None
    assignee_name = ""
    if assignee_id:
        a = db.session.get(User, assignee_id)
        assignee_name = a.display_name if a else ""
    task = Task(
        title=title,
        description=request.form.get("description", "").strip(),
        assignee_id=assignee_id,
        assignee_name=assignee_name,
        due_date=_parse_date(request.form.get("due_date")),
        status=request.form.get("status") or "חדש",
        priority=request.form.get("priority") or "בינונית",
        created_by_id=me.id,
        created_by_name=me.display_name,
    )
    db.session.add(task)
    db.session.commit()
    log_action("פתיחת משימה", title)
    flash("המשימה נוספה בהצלחה", "ok")
    return redirect(url_for("tasks"))


@app.route("/task/<int:task_id>/edit", methods=["POST"])
@login_required
def task_edit(task_id):
    task = db.session.get(Task, task_id)
    if not task:
        abort(404)
    task.title = request.form.get("title", task.title).strip()
    task.description = request.form.get("description", "").strip()
    task.due_date = _parse_date(request.form.get("due_date"))
    task.priority = request.form.get("priority") or task.priority
    assignee_id = request.form.get("assignee_id", type=int) or None
    task.assignee_id = assignee_id
    if assignee_id:
        a = db.session.get(User, assignee_id)
        task.assignee_name = a.display_name if a else ""
    else:
        task.assignee_name = ""
    db.session.commit()
    log_action("עריכת משימה", task.title)
    flash("המשימה עודכנה", "ok")
    return redirect(url_for("tasks"))


@app.route("/task/<int:task_id>/status", methods=["POST"])
@login_required
def task_status(task_id):
    task = db.session.get(Task, task_id)
    if not task:
        abort(404)
    new_status = request.form.get("status", "")
    if new_status in TASK_STATUSES and new_status != task.status:
        task.status = new_status
        db.session.commit()
        log_action("שינוי סטטוס משימה", task.title, f"ל-{new_status}")
        flash(f"סטטוס המשימה שונה ל-{new_status}", "ok")
    return redirect(url_for("tasks"))


@app.route("/task/<int:task_id>/delete", methods=["POST"])
@login_required
def task_delete(task_id):
    task = db.session.get(Task, task_id)
    if task:
        title = task.title
        db.session.delete(task)
        db.session.commit()
        log_action("מחיקת משימה", title)
        flash("המשימה נמחקה", "ok")
    return redirect(url_for("tasks"))


# ---------------------------------------------------------------------------
# ניהול מודולים
# ---------------------------------------------------------------------------
@app.route("/modules", methods=["GET", "POST"])
@login_required
def modules():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        color = request.form.get("color", "#6b7280")
        parent_id = request.form.get("parent_id", type=int) or None
        # תת-מודול יכול להיות רק תחת מודול ראשי (רמה אחת)
        if parent_id:
            parent = db.session.get(Module, parent_id)
            if not parent or parent.parent_id:
                parent_id = None
        if name and not Module.query.filter_by(name=name).first():
            db.session.add(Module(name=name, color=color, parent_id=parent_id))
            db.session.commit()
            log_action("הוספת תת-מודול" if parent_id else "הוספת מודול", name)
            flash("המודול נוסף", "ok")
        else:
            flash("שם מודול ריק או קיים כבר", "error")
        return redirect(url_for("modules"))
    top_modules = Module.query.filter_by(parent_id=None).order_by(Module.name).all()
    links = ModuleLink.query.all()
    return render_template(
        "modules.html", modules=top_modules,
        all_modules=modules_for_select(), links=links,
    )


@app.route("/modules/<int:mod_id>/edit", methods=["POST"])
@login_required
def module_edit(mod_id):
    mod = db.session.get(Module, mod_id)
    if not mod:
        abort(404)
    new_name = request.form.get("name", "").strip()
    new_color = request.form.get("color", mod.color)
    if not new_name:
        flash("שם המודול לא יכול להיות ריק", "error")
        return redirect(url_for("modules"))
    exists = Module.query.filter(Module.name == new_name, Module.id != mod.id).first()
    if exists:
        flash("קיים כבר מודול בשם הזה", "error")
        return redirect(url_for("modules"))
    old_name = mod.name
    mod.name = new_name
    mod.color = new_color
    db.session.commit()
    # השיוכים נשמרים אוטומטית — הקישור הוא לפי מזהה, לא לפי שם
    log_action("עריכת מודול", new_name, f"מ-{old_name}" if old_name != new_name else "")
    flash("המודול עודכן. כל הדרישות המשויכות נשמרו.", "ok")
    return redirect(url_for("modules"))


@app.route("/modules/<int:mod_id>/delete", methods=["POST"])
@login_required
def module_delete(mod_id):
    mod = db.session.get(Module, mod_id)
    if mod:
        mod_name = mod.name
        # תתי-המודולים הופכים לראשיים (לא נמחקים), והדרישות נשארות ללא שיוך
        for child in mod.children:
            child.parent_id = None
        try:
            ModuleLink.query.filter(
                db.or_(ModuleLink.source_id == mod.id, ModuleLink.target_id == mod.id)
            ).delete(synchronize_session=False)
            db.session.delete(mod)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"המחיקה נכשלה: {e}", "error")
            return redirect(url_for("modules"))
        log_action("מחיקת מודול", mod_name)
        flash("המודול נמחק", "ok")
    return redirect(url_for("modules"))


@app.route("/modules/link", methods=["POST"])
@login_required
def module_link_add():
    source_id = request.form.get("source_id", type=int)
    target_id = request.form.get("target_id", type=int)
    label = request.form.get("label", "").strip()
    if not source_id or not target_id or source_id == target_id:
        flash("יש לבחור שני מודולים שונים", "error")
        return redirect(url_for("modules"))
    exists = ModuleLink.query.filter_by(source_id=source_id, target_id=target_id).first()
    if exists:
        flash("הקשר כבר קיים", "error")
        return redirect(url_for("modules"))
    src = db.session.get(Module, source_id)
    tgt = db.session.get(Module, target_id)
    db.session.add(ModuleLink(source_id=source_id, target_id=target_id, label=label))
    db.session.commit()
    log_action("הוספת קשר מודולים", f"{src.name} ← {tgt.name}")
    flash("הקשר נוסף", "ok")
    return redirect(url_for("modules"))


@app.route("/modules/link/<int:link_id>/delete", methods=["POST"])
@login_required
def module_link_delete(link_id):
    link = db.session.get(ModuleLink, link_id)
    if link:
        db.session.delete(link)
        db.session.commit()
        log_action("מחיקת קשר מודולים", f"{link.source.name} ← {link.target.name}")
        flash("הקשר נמחק", "ok")
    return redirect(url_for("modules"))


@app.route("/modules/diagram")
@login_required
def modules_diagram():
    all_modules = Module.query.order_by(Module.parent_id.isnot(None), Module.name).all()
    links = ModuleLink.query.all()
    return render_template("module_diagram.html", all_modules=all_modules, links=links)


# ---------------------------------------------------------------------------
# ניהול משתמשים — אדמין בלבד
# ---------------------------------------------------------------------------
@app.route("/admin")
@admin_required
def admin():
    name_q = request.args.get("name", "").strip()
    phone_q = request.args.get("phone", "").strip()
    role_q = request.args.get("role", "")
    query = User.query
    if name_q:
        query = query.filter(User.display_name.ilike(f"%{name_q}%"))
    if phone_q:
        query = query.filter(User.phone.ilike(f"%{phone_q}%"))
    if role_q == "admin":
        query = query.filter_by(is_admin=True)
    elif role_q == "user":
        query = query.filter_by(is_admin=False)
    users = query.order_by(User.created_at).all()
    return render_template(
        "admin.html", users=users,
        name_q=name_q, phone_q=phone_q, role_q=role_q,
    )


@app.route("/admin/user/add", methods=["POST"])
@admin_required
def admin_user_add():
    username = request.form.get("username", "").strip()
    display_name = request.form.get("display_name", "").strip() or username
    password = request.form.get("password", "")
    phone = request.form.get("phone", "").strip()
    is_admin = bool(request.form.get("is_admin"))

    if not username or not password:
        flash("חובה להזין שם משתמש וסיסמה", "error")
        return redirect(url_for("admin"))
    if User.query.filter_by(username=username).first():
        flash("שם המשתמש כבר קיים", "error")
        return redirect(url_for("admin"))

    user = User(username=username, display_name=display_name, phone=phone, is_admin=is_admin)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    log_action("רישום משתמש", display_name, "אדמין" if is_admin else "משתמש")
    flash(f"המשתמש {display_name} נוסף בהצלחה", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
@admin_required
def admin_user_delete(user_id):
    if user_id == current_user().id:
        flash("לא ניתן למחוק את המשתמש שאיתו את מחוברת", "error")
        return redirect(url_for("admin"))
    user = db.session.get(User, user_id)
    if user:
        user_name = user.display_name
        db.session.delete(user)
        db.session.commit()
        log_action("מחיקת משתמש", user_name)
        flash("המשתמש נמחק", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/user/<int:user_id>/rename", methods=["POST"])
@admin_required
def admin_user_rename(user_id):
    user = db.session.get(User, user_id)
    new_name = request.form.get("display_name", "").strip()
    if user and new_name:
        old_name = user.display_name
        user.display_name = new_name
        db.session.commit()
        log_action("שינוי שם משתמש", new_name, f"מ-{old_name}")
        flash(f"השם עודכן ל-{new_name}", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/user/<int:user_id>/phone", methods=["POST"])
@admin_required
def admin_user_phone(user_id):
    user = db.session.get(User, user_id)
    if user:
        user.phone = request.form.get("phone", "").strip()
        db.session.commit()
        log_action("עדכון טלפון", user.display_name)
        flash(f"הטלפון של {user.display_name} עודכן", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/user/<int:user_id>/reset", methods=["POST"])
@admin_required
def admin_user_reset(user_id):
    user = db.session.get(User, user_id)
    new_password = request.form.get("password", "")
    if user and new_password:
        user.set_password(new_password)
        db.session.commit()
        log_action("איפוס סיסמה", user.display_name)
        flash(f"הסיסמה של {user.display_name} אופסה", "ok")
    return redirect(url_for("admin"))


# ---------------------------------------------------------------------------
# לוג פעולות — אדמין בלבד
# ---------------------------------------------------------------------------
@app.route("/log")
@admin_required
def activity_log():
    user_filter = request.args.get("user", type=int)
    query = ActionLog.query
    if user_filter:
        query = query.filter_by(user_id=user_filter)
    entries = query.order_by(ActionLog.timestamp.desc()).limit(500).all()
    users = User.query.order_by(User.display_name).all()
    # סימון שהלוג נצפה — כדי לאפס את הנקודה האדומה
    me = current_user()
    me.log_seen_at = datetime.utcnow()
    db.session.commit()
    return render_template(
        "log.html", entries=entries, users=users, user_filter=user_filter
    )


# ---------------------------------------------------------------------------
# אתחול בסיס הנתונים ויצירת אדמין ראשוני
# ---------------------------------------------------------------------------
def ensure_schema():
    """
    מיגרציה קלה ובטוחה: מוסיפה עמודות חדשות לטבלאות קיימות בלבד.
    מבצעת אך ורק ALTER TABLE ... ADD COLUMN — פעולה שלעולם אינה מוחקת
    או משנה נתונים קיימים. רצה בכל עלייה ומדלגת על מה שכבר קיים.
    """
    from sqlalchemy import inspect, text
    insp = inspect(db.engine)

    # (טבלה, עמודה, סוג) — כל אלה תוספות בלבד
    needed = [
        ("user", "phone", "VARCHAR(40)"),
        ("user", "log_seen_at", "TIMESTAMP"),
        ("bug", "assignee_id", "INTEGER"),
        ("bug", "assignee_name", "VARCHAR(120)"),
        ("notification", "bug_id", "INTEGER"),
        ("requirement", "implementer_id", "INTEGER"),
        ("requirement", "implementer_name", "VARCHAR(120)"),
        ("bug", "priority", "VARCHAR(40)"),
        ("task", "priority", "VARCHAR(40)"),
        ("module", "parent_id", "INTEGER"),
        ("requirement", "actual_implementer_id", "INTEGER"),
        ("requirement", "actual_implementer_name", "VARCHAR(120)"),
    ]
    for table, column, coltype in needed:
        try:
            existing = [c["name"] for c in insp.get_columns(table)]
        except Exception:
            continue  # הטבלה עוד לא קיימת — create_all ייצור אותה מלאה
        if column not in existing:
            try:
                db.session.execute(
                    text(f'ALTER TABLE "{table}" ADD COLUMN {column} {coltype}')
                )
                db.session.commit()
                print(f"[migrate] נוספה עמודה {column} לטבלת {table}")
            except Exception as e:
                db.session.rollback()
                print(f"[migrate] דילוג על {table}.{column}: {e}")


def migrate_statuses():
    """
    איחוד סטטוסים לשלושה משותפים (חדש / בתהליך / בוצע) בכל שלושת הסוגים,
    ומילוי עדיפות ברירת מחדל 'בינונית'. שומר על הנתונים — רק מעדכן ערכים.
    אידמפוטנטי (אפשר להריץ בכל עלייה).
    """
    from sqlalchemy import text
    mappings = {
        "requirement": {"הצעה": "חדש", "בבדיקה": "בתהליך", "אושרה": "בתהליך",
                         "בפיתוח": "בתהליך", "הושלמה": "בוצע", "נדחתה": "בוצע"},
        "bug":         {"פתוח": "חדש", "סגור": "בוצע"},
        "task":        {"לביצוע": "חדש", "בתהליך": "בתהליך", "הושלם": "בוצע"},
    }
    for table, m in mappings.items():
        try:
            for old, new in m.items():
                db.session.execute(
                    text(f"UPDATE \"{table}\" SET status=:new WHERE status=:old"),
                    {"new": new, "old": old},
                )
            # כל ערך שאינו אחד מהשלושה החדשים → 'חדש'
            db.session.execute(text(
                f"UPDATE \"{table}\" SET status='חדש' "
                f"WHERE status IS NULL OR status NOT IN ('חדש','בתהליך','בוצע')"
            ))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[migrate] דילוג על איחוד סטטוס ב-{table}: {e}")
    # מילוי עדיפות ריקה בבאגים ומשימות
    for table in ("bug", "task"):
        try:
            db.session.execute(text(
                f"UPDATE \"{table}\" SET priority='בינונית' WHERE priority IS NULL OR priority=''"
            ))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[migrate] דילוג על מילוי עדיפות ב-{table}: {e}")


def init_db():
    with app.app_context():
        db.create_all()
        ensure_schema()
        migrate_statuses()
        if not User.query.first():
            admin_username = os.environ.get("ADMIN_USERNAME", "dikla")
            admin_password = os.environ.get("ADMIN_PASSWORD", "changeme123")
            admin_name = os.environ.get("ADMIN_NAME", "דקלה")
            admin_user = User(
                username=admin_username,
                display_name=admin_name,
                is_admin=True,
            )
            admin_user.set_password(admin_password)
            db.session.add(admin_user)
            db.session.commit()
            print(f"[init] נוצר משתמש אדמין: {admin_username}")


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
